from fastapi import BackgroundTasks, FastAPI, File, UploadFile, HTTPException, Form, Depends, Query
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import shutil
from typing import List, Dict, Optional, Union, Any
from pydantic import BaseModel, Field
import pandas as pd
import json
import os
from openai import OpenAI
from docx import Document
import tempfile
from pathlib import Path
import openpyxl
from dotenv import load_dotenv
import threading
import xml.etree.ElementTree as ET
import uvicorn
from openpyxl import load_workbook
import logging
from sqlalchemy.orm import Session
from datetime import datetime
# from scripts.anonymise import process_file_for_anonymization, de_anonymize_with_generated_terms

# Load environment variables
load_dotenv()


# Capstone (infal round for complete demo):
# -> UI changes/functions
# Display document/module name e.g SCR_001
# Refine default template to include 1. dynamically naming after project + module + document
# Add option to create versions of end to end test cases (modules combined to one, add all mandatory test specs + targeted tests)
 
# Data Needed:
# -> check for whether SCR affects mandatory TC
# -> if SCR, generate comparison with mandatory TC, let user edit and save to DB. Changes are categorised as targeted
 


# Initialize OpenAI API client
openai = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))  # Replace with your API key

# Initialize FastAPI app
app = FastAPI()
# Determine environment
CONFIG_MODE = os.getenv("CONFIG_MODE", "local")  # Default to 'local'
BASE_DIR = Path(__file__).resolve().parent

# Define file paths based on environment
if CONFIG_MODE == "docker":
    STATIC_DIR = BASE_DIR / "templates"
    TEMPLATE_PATH = BASE_DIR / "testdata" / "Default_Template.xlsx"
    OUTPUT_PATH = BASE_DIR / "generateddata" / "updated_test_cases.xlsx"
    from app.db.database import SessionLocal, engine, Base
    from app.db.models import Project, Module, TestScripts, Suite
    print(TEMPLATE_PATH)
else:
    STATIC_DIR = BASE_DIR / "templates"
    TEMPLATE_PATH = BASE_DIR / "testdata" / "Default_Template.xlsx"
    OUTPUT_PATH = BASE_DIR / "generateddata" / "updated_test_cases.xlsx"
    from db.database import SessionLocal, engine, Base
    from db.models import Project, TestScripts
    
print(f"Running in {CONFIG_MODE} mode")

logging.basicConfig(level=logging.DEBUG)

app.mount("/frontend", StaticFiles(directory="../frontend/dist", html=True), name="static")


# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allow all methods
    allow_headers=["*"],  # Allow all headers
)

# DB starts here

Base.metadata.create_all(bind=engine)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()
        
        
class ProjectCreate(BaseModel):
    name: str
    description: str = ""

class ProjectResponse(BaseModel):
    id: int
    name: str
    description: str
    created_at: datetime
    updated_at: datetime

    class Config:
        orm_mode = True
        
@app.get("/projects", response_model=list[ProjectResponse])
def read_projects(db: Session = Depends(get_db)):
    projects = db.query(Project).all()
    return projects


### Fixed: POST `/projects` ###
@app.post("/projects", response_model=ProjectResponse)
def create_project(project: ProjectCreate, db: Session = Depends(get_db)):
    try:
        new_project = Project(
            name=project.name,
            description=project.description,
        )

        db.add(new_project)
        db.commit()
        db.refresh(new_project)

        return new_project

    except Exception as e:
        db.rollback()
        logging.error(f"Error creating project: {e}")
        raise HTTPException(status_code=500, detail="Error creating project")
    

@app.get("/projects/{project_id}", response_model=ProjectResponse)
def get_project(project_id: int, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.id == project_id).first()
    
    # Debugging output
    if not project:
        logging.error(f"Project with ID {project_id} not found in the database.")
        raise HTTPException(status_code=404, detail="Project not found")
    
    logging.info(f"Project found: {project.name}")
    return project

@app.delete("/projects/{project_id}")
def delete_project(project_id: int, db: Session = Depends(get_db)):
    # Fetch the project
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Delete associated modules and their test scripts
    for module in project.modules:
        db.query(TestScripts).filter(TestScripts.module_id == module.id).delete()
        db.delete(module)

    # Delete associated suites
    for suite in project.suites:
        db.delete(suite)

    # Finally, delete the project itself
    db.delete(project)
    db.commit()

    return {"message": "Project and all associated data deleted successfully"}

#modules

class ModuleCreate(BaseModel):
    project_id: int
    name: str
    description: str = ""  # Default to empty if not provided

class ModuleResponse(BaseModel):
    id: int
    project_id: int
    project_specific_id: int
    name: str
    description: str
    created_at: datetime
    updated_at: datetime
    script_content: str
    scr_update: str

    class Config:
        orm_mode = True


### GET all modules ###
@app.get("/modules", response_model=list[ModuleResponse])
def read_modules(db: Session = Depends(get_db)):
    modules = db.query(Module).all()
    return modules

### GET modules by `project_id` ###
@app.get("/modules/{project_id}", response_model=list[ModuleResponse])
def get_module_by_project_id(project_id: int, db: Session = Depends(get_db)):
    module = db.query(Module).filter(
        Module.project_id == project_id,
    ).all()

    if not module:
        logging.error(f"project {project_id} not found.")
        raise HTTPException(status_code=404, detail="Unknown module in the specified project")
    
    return module


### GET a module by its global `id` ###
@app.get("/modules/{module_id}", response_model=ModuleResponse)
def get_module(module_id: int, db: Session = Depends(get_db)):
    module = db.query(Module).filter(Module.id == module_id).first()
    
    if not module:
        logging.error(f"Module with ID {module_id} not found.")
        raise HTTPException(status_code=404, detail="Module not found")
    
    return module


@app.get("/modules/{project_id}/{project_specific_id}", response_model=ModuleResponse)
def get_module_by_project_specific_id(
    project_id: int, 
    project_specific_id: int, 
    db: Session = Depends(get_db)
):
    module = db.query(Module).filter(
        Module.project_id == project_id,
        Module.project_specific_id == project_specific_id
    ).first()

    if not module:
        logging.error(f"Module {project_specific_id} not found in project {project_id}.")
        raise HTTPException(status_code=404, detail="Module not found in the specified project")
    
    return module


### POST: Create a new module (auto-generates `project_specific_id`) ###
@app.post("/modules", response_model=ModuleResponse)
def create_module(module: ModuleCreate, db: Session = Depends(get_db)):
    try:
        # Get the latest `project_specific_id` for the given project
        latest_module = (
            db.query(Module)
            .filter(Module.project_id == module.project_id)
            .order_by(Module.project_specific_id.desc())
            .first()
        )
        new_project_specific_id = (latest_module.project_specific_id + 1) if latest_module else 1

        new_module = Module(
            project_id=module.project_id,
            project_specific_id=new_project_specific_id,
            name=module.name,
            description=module.description,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
            script_content="",
            scr_update=""
        )

        db.add(new_module)
        db.commit()
        db.refresh(new_module)
        return new_module

    except Exception as e:
        db.rollback()
        logging.error(f"Error creating module: {e}")
        raise HTTPException(status_code=500, detail="Error creating module")
    
    
class ModuleUpdateRequest(BaseModel):
    script_content: List[dict]  # Assuming script_content is a list of test cases

    class Config:
        orm_mode = True
        
    
@app.put("/modules/{project_id}/{project_specific_id}/update_mtc", response_model=dict)
def update_module_script_content(
    project_id: int,
    project_specific_id: int,
    request: ModuleUpdateRequest,
    db: Session = Depends(get_db),
):
    # Fetch the module
    module = db.query(Module).filter(
        Module.project_id == project_id,
        Module.project_specific_id == project_specific_id
    ).first()

    if not module:
        logging.error(f"Module {project_specific_id} not found in project {project_id}.")
        raise HTTPException(status_code=404, detail="Module not found in the specified project")

    try:
        # Convert script_content to JSON format
        module.script_content = json.dumps(request.script_content)  # Convert Python object to JSON string

        # Commit changes
        db.commit()
        db.refresh(module)

        return {"message": "Module script_content updated successfully!"}
    
    except Exception as e:
        db.rollback()
        logging.error(f"Database error: {e}")
        raise HTTPException(status_code=500, detail="Database error while updating module")
    
    
@app.put("/modules/{project_id}/{project_specific_id}/update_ttc", response_model=dict)
def update_module_script_content(
    project_id: int,
    project_specific_id: int,
    request: ModuleUpdateRequest,
    db: Session = Depends(get_db),
):
    # Fetch the module
    module = db.query(Module).filter(
        Module.project_id == project_id,
        Module.project_specific_id == project_specific_id
    ).first()

    if not module:
        logging.error(f"Module {project_specific_id} not found in project {project_id}.")
        raise HTTPException(status_code=404, detail="Module not found in the specified project")

    try:
        # Convert script_content to JSON format
        module.scr_update = json.dumps(request.script_content)  # Convert Python object to JSON string

        # Commit changes
        db.commit()
        db.refresh(module)

        return {"message": "Module script_content updated successfully!"}
    
    except Exception as e:
        db.rollback()
        logging.error(f"Database error: {e}")
        raise HTTPException(status_code=500, detail="Database error while updating module")
    
    
## Suites

class SuiteResponse(BaseModel):
    id: int
    project_id: int
    project_specific_id: int
    name: str
    description: str
    created_at: datetime
    updated_at: datetime
    script_content: str

    class Config:
        orm_mode = True


### GET suites by `project_id` ###
@app.get("/suites/{project_id}", response_model=list[SuiteResponse])
def get_suites_by_project_id(project_id: int, db: Session = Depends(get_db)):
    suite = db.query(Suite).filter(
        Suite.project_id == project_id,
    ).all()

    if not suite:
        logging.error(f"project {project_id} not found.")
        raise HTTPException(status_code=404, detail="Unknown suite in the specified project")
    
    return suite

### POST: Create a new module (auto-generates `project_specific_id`) ###
@app.post("/suites", response_model=SuiteResponse)
def create_suite(module: ModuleCreate, db: Session = Depends(get_db)):
    try:
        # Get the latest `project_specific_id` for the given project
        latest_module = (
            db.query(Suite)
            .filter(Suite.project_id == module.project_id)
            .order_by(Suite.project_specific_id.desc())
            .first()
        )
        new_project_specific_id = (latest_module.project_specific_id + 1) if latest_module else 1

        new_module = Suite(
            project_id=module.project_id,
            project_specific_id=new_project_specific_id,
            name=module.name,
            description=module.description,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
            script_content="",
        )

        db.add(new_module)
        db.commit()
        db.refresh(new_module)
        return new_module

    except Exception as e:
        db.rollback()
        logging.error(f"Error creating module: {e}")
        raise HTTPException(status_code=500, detail="Error creating module")
    
    
class ModuleUpdateRequest(BaseModel):
    script_content: List[dict]  # Assuming script_content is a list of test cases

    class Config:
        orm_mode = True
        
        

@app.get("/search", response_model=List[dict])
def search_items(
    query: str = Query(None, min_length=1),
    db: Session = Depends(get_db)
):
    if not query:
        return []
    query = f"%{query}%"  # SQL LIKE pattern for flexible search
    # Fetch modules
    modules = db.query(Module).filter(Module.name.ilike(query)).all()
    module_results = [
        {
            "name": module.name,
            "type": "Module",
            "link": f"/preview/{module.project.name}/{module.project_id}/{module.name}/{module.id}"
        }
        for module in modules
    ]
    # Fetch suites
    suites = db.query(Suite).filter(Suite.name.ilike(query)).all()
    suite_results = [
        {
            "name": suite.name,
            "type": "Suite",
            "link": f"/project-test-suite/{suite.project.name}/{suite.project_id}/{suite.name}/{suite.id}"
        }
        for suite in suites
    ]
    return module_results + suite_results

# TCs

class TestScriptCreate(BaseModel):
    id: int
    module_id: int
    module_specific_id: int
    name: str
    description: str
    created_at: datetime
    updated_at: datetime
    script_content: str

    class Config:
        orm_mode = True
        
        
class TestScriptResponse(BaseModel):
    id: int
    module_id: int
    module_specific_id: int
    name: str
    description: str
    created_at: datetime
    updated_at: datetime
    script_content: str

    class Config:
        orm_mode = True
    
    
### POST: Create a new ts (auto-generates `project_specific_id`) ###
@app.post("/test-scripts", response_model=TestScriptResponse)
def create_test_scripts(module: TestScriptCreate, db: Session = Depends(get_db)):
    try:
        # Get the latest `project_specific_id` for the given project
        latest_module = (
            db.query(TestScripts)
            .filter(TestScripts.module_id == module.id)
            .order_by(TestScripts.module_specific_id.desc())
            .first()
        )
        new_module_specific_id = (latest_module.module_specific_id + 1) if latest_module else 1

        new_module = TestScripts(
            module_id=module.id,
            module_specific_id=new_module_specific_id,
            name=module.name,
            description=module.description,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
            script_content=""
        )

        db.add(new_module)
        db.commit()
        db.refresh(new_module)
        return new_module

    except Exception as e:
        db.rollback()
        logging.error(f"Error creating test script: {e}")
        raise HTTPException(status_code=500, detail="Error creating test script")
    
    
@app.get("/test-scripts/{module_id}/{module_specific_id}", response_model=TestScriptResponse)
def get_test_script_by_project_specific_id(
    project_id: int, 
    project_specific_id: int, 
    db: Session = Depends(get_db)
):
    module = db.query(TestScripts).filter(
        TestScripts.project_id == project_id,
        TestScripts.project_specific_id == project_specific_id
    ).first()

    if not module:
        logging.error(f"Test Script {project_specific_id} not found in project {project_id}.")
        raise HTTPException(status_code=404, detail="Test Script not found in the specified project")
    
    return module
    
    
@app.put("/test-scripts/{project_id}/{project_specific_id}", response_model=dict)
def update_module_script_content(
    project_id: int,
    project_specific_id: int,
    request: ModuleUpdateRequest,
    db: Session = Depends(get_db),
):
    # Fetch the module
    module = db.query(TestScripts).filter(
        Module.project_id == project_id,
        Module.project_specific_id == project_specific_id
    ).first()

    if not module:
        logging.error(f"Test Script {project_specific_id} not found in project {project_id}.")
        raise HTTPException(status_code=404, detail="Module not found in the specified project")

    try:
        # Convert script_content to JSON format
        module.script_content = json.dumps(request.script_content)  # Convert Python object to JSON string

        # Commit changes
        db.commit()
        db.refresh(module)

        return {"message": "Module script_content updated successfully!"}
    
    except Exception as e:
        db.rollback()
        logging.error(f"Database error: {e}")
        raise HTTPException(status_code=500, detail="Database error while updating module")


# rest of app routes

@app.get("/", response_class=HTMLResponse)
async def get_ui():
    with open(BASE_DIR / "templates" / "index.html") as f:
        return HTMLResponse(content=f.read())


def cleanup_files(temp_dir, file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
    if os.path.exists(temp_dir):
        os.rmdir(temp_dir)


def cleanup_temp_dir(temp_dir):
    try:
        shutil.rmtree(temp_dir)  # Removes the directory and all its contents
        print(f"Cleaned up temporary directory: {temp_dir}")
    except Exception as e:
        print(f"Failed to clean up temporary directory: {e}")
        
# Function to load JSON files
def load_request_data(file_path):
    with open(file_path, "r") as file:
        return json.load(file)

# Function to run the FastAPI server
def run_fastapi():
    uvicorn.run(app, host="127.0.0.1", port=8000)

@app.post("/upload/")
async def upload_file(file: UploadFile = File(...)):
    print("upload backend triggered")
    try:
        # Save uploaded file to temporary directory
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, file.filename)
        test_scripts = ""
        with open(file_path, "wb") as f:
            f.write(await file.read())
            
        if (file.filename.split('.').pop().lower() == "docx"):
            # Extract text from the Word document
            document_text = extract_text_from_docx(file_path)

            # Generate test scripts using OpenAI
            test_scripts = await generate_test_scripts(document_text)
        else:
            # Extract text from the Excel file
            document_text = extract_excel_content(file_path)

            # Generate test scripts using OpenAI
            test_scripts = await generate_test_scripts(document_text)

        # Generate Excel file with test scripts
        excel_file_path = template_excel(test_scripts, TEMPLATE_PATH, OUTPUT_PATH)

        # Send the Excel file as a response
        return FileResponse(excel_file_path, filename="test-scripts.xlsx")
    finally:
        # Clean up temporary files
        if os.path.exists(file_path):
            os.remove(file_path)
            

@app.post("/generate-test-cases")
async def generate_test_cases(file: UploadFile = File(...)):
    try:
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, file.filename)
        with open(file_path, "wb") as f:
            f.write(await file.read())
        
        if file.filename.split('.').pop().lower() == "docx":
            document_text = extract_text_from_docx(file_path)
        else:
            document_text = extract_excel_content(file_path)
        
        test_cases = await generate_test_scripts(document_text)
        return {"test_cases": test_cases}
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)
            
# @app.post("/generate-full-flow")
# async def generate_full_flow(selectedModules: string[], testStrategy: string):
#     try:
#         temp_dir = tempfile.mkdtemp()
#         file_path = os.path.join(temp_dir, file.filename)
        
#         test_cases = await generate_test_scripts(document_text)
#         return {"test_cases": test_cases}
#     finally:
#         if os.path.exists(file_path):
#             os.remove(file_path)
            

class TestCase(BaseModel):
    Test_Case_ID: str
    Test_Case_Name: str
    Test_Case_Type: str
    Pre_Condition: str
    Actor_s: str
    Test_Data: str
    Step_Description: str
    Expected_Result: str

class GenerateExcelRequest(BaseModel):
    test_cases: List[TestCase]


@app.post("/generate-excel")
async def generate_excel(request: GenerateExcelRequest):
    test_cases = request.test_cases
    print("Received test cases:", test_cases)  # Log the parsed test cases
    excel_file_path = template_excel(test_cases, TEMPLATE_PATH, OUTPUT_PATH)
    return FileResponse(
        excel_file_path,
        headers={"Content-Disposition": 'attachment; filename="test-scripts.xlsx"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Extract text from a Word document
def extract_text_from_docx(file_path: str) -> str:
    doc = Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs if para.text.strip()])

def extract_excel_content(file_path):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        output = []

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            output.append(f"\n### Sheet: {sheet_name} ###\n")

            # Iterate through rows and columns
            for row in sheet.iter_rows(values_only=True):
                row_content = "\t".join(str(cell) if cell is not None else "" for cell in row)
                output.append(row_content)

        return "\n".join(output)

    except Exception as e:
        return f"Error reading Excel file: {str(e)}"

# Query OpenAI to generate test scripts
async def generate_test_scripts(document_text: str) -> list:
    response = openai.chat.completions.create(
        model="gpt-4o-mini",  # Replace with your model
        messages=[
            {"role": "system", "content": "You are an assistant that generates software test scripts."},
            {"role": "user", "content": (
            f"Generate test scripts for the following document:\n\n{document_text}. "
            "Limit your response to only the test script contents in JSON format. "
            "Return the test cases as a JSON array (list) with each test case containing the following attributes: "
            "Test_Case_ID, Test_Case_Name, Test_Case_Type, Pre_Condition, Actor_s, Test_Data, Step_Description, and Expected_Result. "
            "Test_Case_Type values are limited to: Mandatory, SCR Change, Boundary Test, Edge Case, Exception Test. For each test case, select the most appropriate Test_Case_Type value."
            "Do not include any additional text, just the JSON array in text. Every attribute should be a string. If a test case read seems appropriate, you may append it to your output as-is."
            )}
        ]
    )
    output = response.choices[0].message.content
    
    # Remove the ```json and ``` delimiters if present
    if output.startswith("```json"):
        output = output[7:]  # Remove the initial ```json
    if output.endswith("```"):
        output = output[:-3]  # Remove the closing ```
    
    print(output)
    
    try:
        # Parse the JSON response
        test_cases = json.loads(output)
        return test_cases  # Return the list of test case dictionaries
    except json.JSONDecodeError:
        raise ValueError("The API response is not valid JSON. Ensure the prompt enforces proper JSON formatting.")
    

# Query OpenAI to generate test scripts
async def generate_test_suites_full_flow(test_cases_json: str, instructions: str) -> list:
    response = openai.chat.completions.create(
        model="gpt-4o-mini",  # Replace with your model
        messages=[
            {"role": "system", "content": "You are an assistant that generates software test scripts."},
            {"role": "user", "content": (
            f"Generate test scripts for the following document:\n\n{test_cases_json}. "
            "Limit your response to only the test script contents in JSON format. "
            "Return the test cases as a JSON array (list) with each test case containing the following attributes: "
            "Test_Case_ID, Test_Case_Name, Test_Case_Type, Pre_Condition, Actor_s, Test_Data, Step_Description, and Expected_Result. "
            "Test_Case_Type values are limited to: Mandatory, SCR Change, Boundary Test, Edge Case, Exception Test. For each test case, select the most appropriate Test_Case_Type value."
            "Do not include any additional text, just the JSON array in text. Every attribute should be a string. If a test case read seems appropriate, you may append it to your output as-is."
            )}
        ]
    )
    output = response.choices[0].message.content
    
    # Remove the ```json and ``` delimiters if present
    if output.startswith("```json"):
        output = output[7:]  # Remove the initial ```json
    if output.endswith("```"):
        output = output[:-3]  # Remove the closing ```
    
    print(output)
    
    try:
        # Parse the JSON response
        test_cases = json.loads(output)
        return test_cases  # Return the list of test case dictionaries
    except json.JSONDecodeError:
        raise ValueError("The API response is not valid JSON. Ensure the prompt enforces proper JSON formatting.")


def template_excel(test_cases: List[TestCase], template_path: str, output_path: str) -> str:
    workbook = load_workbook(template_path)
    sheet = workbook.active

    for row, test_case in enumerate(test_cases, start=28):
        sheet.cell(row=row, column=1).value = test_case.Test_Case_ID
        sheet.cell(row=row, column=2).value = test_case.Test_Case_Name
        sheet.cell(row=row, column=3).value = test_case.Pre_Condition
        sheet.cell(row=row, column=4).value = test_case.Actor_s
        sheet.cell(row=row, column=5).value = str(test_case.Test_Data)
        sheet.cell(row=row, column=6).value = test_case.Step_Description  # Join list into a string
        sheet.cell(row=row, column=7).value = test_case.Expected_Result
    
    # Save the updated workbook to the output path
    temp_dir = tempfile.mkdtemp()
    excel_file_path = os.path.join(temp_dir, "test-scripts.xlsx")
    workbook.save(excel_file_path)
    return excel_file_path

# Ranorex integration

# Pydantic model to validate incoming request data
class AutomationData(BaseModel):
    suite: str
    projectName: str
    websiteUrl: str
    browser: str
    testCases: list

# Helper function to call OpenAI API for generating Ranorex execution code
async def generate_execution_code(test_cases: list) -> str:
    response = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "You are an assistant that generates Ranorex test automation scripts in C#."},
            {"role": "user", "content": (
                "Generate Ranorex-compatible C# test scripts based on the following test cases:\n\n"
                f"{json.dumps(test_cases, indent=2)}\n\n"
                "For each test case, generate a class with methods for each step. Include necessary Ranorex namespaces."
            )}
        ]
    )
    
    execution_code = response.choices[0].message.content
    return execution_code


def setup_ranorex_project(data: AutomationData, execution_code: str):
    # Define base paths
    base_path = os.path.join(os.path.dirname(__file__), "RanorexProjects", data.projectName)
    modules_path = os.path.join(base_path, "Modules")
    repository_path = os.path.join(base_path, "Repository")
    test_suite_path = os.path.join(base_path, "TestSuite")
    reports_path = os.path.join(base_path, "Reports")
    binaries_path = os.path.join(base_path, "Binaries")

    # Create directories if they don't exist
    os.makedirs(modules_path, exist_ok=True)
    os.makedirs(repository_path, exist_ok=True)
    os.makedirs(test_suite_path, exist_ok=True)
    os.makedirs(reports_path, exist_ok=True)
    os.makedirs(binaries_path, exist_ok=True)

    # Create Project.rxsln (Solution File)
    solution_file_path = os.path.join(base_path, f"{data.projectName}.rxsln")
    with open(solution_file_path, 'w') as solution_file:
        solution_file.write(f"""<?xml version="1.0" encoding="utf-8"?>
        <Solution xmlns="http://schemas.ranorex.com/solution">
        <Project>{data.projectName}.rxproj</Project>
        <Repository>{data.projectName}.rxrep</Repository>
        <TestSuite>{data.projectName}.rxtst</TestSuite>
        </Solution>""")

    # Create Project.rxproj (Project File)
    project_file_path = os.path.join(base_path, f"{data.projectName}.rxproj")
    with open(project_file_path, 'w') as project_file:
        project_file.write(f"""<?xml version="1.0" encoding="utf-8"?>
        <Project name="{data.projectName}" browser="{data.browser}" url="{data.websiteUrl}">
        <Modules>
            <Module>{data.projectName}_Tests.cs</Module>
        </Modules>
        </Project>""")

    # Create Project.rxrep (Repository File)
    repository_file_path = os.path.join(repository_path, f"{data.projectName}.rxrep")
    repository_root = ET.Element("Repository")
    for element in data.ui_elements:  # Assuming data.ui_elements is a list of UI elements
        ui_element = ET.SubElement(repository_root, "Element", attrib={
            "name": element["name"],
            "id": element["id"],
            "xpath": element["xpath"]
        })
    repository_tree = ET.ElementTree(repository_root)
    repository_tree.write(repository_file_path, encoding="utf-8", xml_declaration=True)

    # Create TestSuite.rxtst (Test Suite File)
    test_suite_file_path = os.path.join(test_suite_path, f"{data.projectName}.rxtst")
    test_suite_root = ET.Element("TestSuite", attrib={"name": data.suite})
    for test_case in data.test_cases:  # Assuming data.test_cases is a list of test cases
        test_case_element = ET.SubElement(test_suite_root, "TestCase", attrib={"name": test_case["name"]})
        for step in test_case["steps"]:
            step_element = ET.SubElement(test_case_element, "Step", attrib={
                "action": step["action"],
                "element": step["element"],
                "value": step["value"]
            })
    test_suite_tree = ET.ElementTree(test_suite_root)
    test_suite_tree.write(test_suite_file_path, encoding="utf-8", xml_declaration=True)

    # Write the generated execution code to a C# file
    execution_code_path = os.path.join(modules_path, f"{data.projectName}_Tests.cs")
    with open(execution_code_path, 'w') as script_file:
        script_file.write(execution_code)

    print(f"Ranorex project '{data.projectName}' has been successfully created at '{base_path}'.")


# FastAPI endpoint to handle automation setup
@app.post("/setup-automation")
async def setup_automation(automation_data: AutomationData):
    try:
        # Generate execution code from test cases
        execution_code = await generate_execution_code(automation_data.testCases)
        
        # Setup the Ranorex project with generated files
        setup_ranorex_project(automation_data, execution_code)

        return {"message": "Ranorex project setup successfully!", "projectName": automation_data.projectName}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to set up Ranorex project: {str(e)}")
